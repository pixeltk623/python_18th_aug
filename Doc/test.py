# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('C:\\Users\\Sharvan\\Desktop\\excel.xlsx')

# create the sheet object
sheet = wb.active


def excel():
	
	# resize the width of columns in
	# excel spreadsheet
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50

	# write given data to an excel spreadsheet
	# at particular location
	sheet.cell(row=1, column=1).value = "Name"
	sheet.cell(row=1, column=2).value = "Course"
	sheet.cell(row=1, column=3).value = "Semester"
	sheet.cell(row=1, column=4).value = "Form Number"
	sheet.cell(row=1, column=5).value = "Contact Number"
	sheet.cell(row=1, column=6).value = "Email id"
	sheet.cell(row=1, column=7).value = "Address"


# Function to set focus (cursor)
def focus1(event):
	# set focus on the course_field box
	course_field.focus_set()


# Function to set focus
def focus2(event):
	# set focus on the sem_field box
	sem_field.focus_set()


# Function to set focus
def focus3(event):
	# set focus on the form_no_field box
	form_no_field.focus_set()


# Function to set focus
def focus4(event):
	# set focus on the contact_no_field box
	contact_no_field.focus_set()


# Function to set focus
def focus5(event):
	# set focus on the email_id_field box
	email_id_field.focus_set()


# Function to set focus
def focus6(event):
	# set focus on the address_field box
	address_field.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():
	
	# clear the content of text entry box
	name_field.delete(0, END)
	course_field.delete(0, END)
	sem_field.delete(0, END)
	form_no_field.delete(0, END)
	contact_no_field.delete(0, END)
	email_id_field.delete(0, END)
	address_field.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert():
	
	# if user not fill any entry
	# then print "empty input"
	if (name_field.get() == "" and
		course_field.get() == "" and
		sem_field.get() == "" and
		form_no_field.get() == "" and
		contact_no_field.get() == "" and
		email_id_field.get() == "" and
		address_field.get() == ""):
			
		print("empty input")

	else:

		# assigning the max row and max column
		# value upto which data is written
		# in an excel sheet to the variable
		current_row = sheet.max_row
		current_column = sheet.max_column

		# get method returns current text
		# as string which we write into
		# excel spreadsheet at particular location
		sheet.cell(row=current_row + 1, column=1).value = name_field.get()
		sheet.cell(row=current_row + 1, column=2).value = course_field.get()
		sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
		sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
		sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
		sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
		sheet.cell(row=current_row + 1, column=7).value = address_field.get()

		# save the file
		wb.save('C:\\Users\\Sharvan\\Desktop\\excel.xlsx')

		# set focus on the name_field box
		name_field.focus_set()

		# call the clear() function
		clear()


# Driver code
if __name__ == "__main__":
	
	# create a GUI window
	root = Tk()

	# set the background colour of GUI window
	root.configure(background='light green')

	# set the title of GUI window
	root.title("registration form")

	# set the configuration of GUI window
	root.geometry("500x300")

	excel()

	# create a Form label
	heading = Label(root, text="Form", bg="light green")

	# create a Name label
	name = Label(root, text="Name", bg="light green")

	# create a Course label
	course = Label(root, text="Course", bg="light green")

	# create a Semester label
	sem = Label(root, text="Semester", bg="light green")

	# create a Form No. lable
	form_no = Label(root, text="Form No.", bg="light green")

	# create a Contact No. label
	contact_no = Label(root, text="Contact No.", bg="light green")

	# create a Email id label
	email_id = Label(root, text="Email id", bg="light green")

	# create a address label
	address = Label(root, text="Address", bg="light green")

	# grid method is used for placing
	# the widgets at respective positions
	# in table like structure .
	heading.grid(row=0, column=1)
	name.grid(row=1, column=0)
	course.grid(row=2, column=0)
	sem.grid(row=3, column=0)
	form_no.grid(row=4, column=0)
	contact_no.grid(row=5, column=0)
	email_id.grid(row=6, column=0)
	address.grid(row=7, column=0)

	# create a text entry box
	# for typing the information
	name_field = Entry(root)
	course_field = Entry(root)
	sem_field = Entry(root)
	form_no_field = Entry(root)
	contact_no_field = Entry(root)
	email_id_field = Entry(root)
	address_field = Entry(root)

	# bind method of widget is used for
	# the binding the function with the events

	# whenever the enter key is pressed
	# then call the focus1 function
	name_field.bind("<Return>", focus1)

	# whenever the enter key is pressed
	# then call the focus2 function
	course_field.bind("<Return>", focus2)

	# whenever the enter key is pressed
	# then call the focus3 function
	sem_field.bind("<Return>", focus3)

	# whenever the enter key is pressed
	# then call the focus4 function
	form_no_field.bind("<Return>", focus4)

	# whenever the enter key is pressed
	# then call the focus5 function
	contact_no_field.bind("<Return>", focus5)

	# whenever the enter key is pressed
	# then call the focus6 function
	email_id_field.bind("<Return>", focus6)

	# grid method is used for placing
	# the widgets at respective positions
	# in table like structure .
	name_field.grid(row=1, column=1, ipadx="100")
	course_field.grid(row=2, column=1, ipadx="100")
	sem_field.grid(row=3, column=1, ipadx="100")
	form_no_field.grid(row=4, column=1, ipadx="100")
	contact_no_field.grid(row=5, column=1, ipadx="100")
	email_id_field.grid(row=6, column=1, ipadx="100")
	address_field.grid(row=7, column=1, ipadx="100")

	# call excel function
	excel()

	# create a Submit Button and place into the root window
	submit = Button(root, text="Submit", fg="Black",
							bg="Red", command=insert)
	submit.grid(row=8, column=1)

	# start the GUI
	root.mainloop()

'''pass is just a placeholder for
functionality to be added later.'''
# sequence = {'p', 'a', 's', 's'}
# for val in sequence:
#     print(val)

# Use of break statement inside the loop

# for val in "string":
#     if val == "i":
#         break
#     print(val)

# print("The end")






# # Python for Loop
# # In this article, you'll learn to iterate over a sequence of elements using the different variations of for loop.

# # Video: Python for Loop

# # What is for loop in Python?
# # The for loop in Python is used to iterate over a sequence (list, tuple, string) or other iterable objects. Iterating over a sequence is called traversal.

# # Syntax of for Loop

# # for val in sequence:
# #     loop body
# # Here, val is the variable that takes the value of the item inside the sequence on each iteration.

# # Loop continues until we reach the last item in the sequence. The body of for loop is separated from the rest of the code using indentation.

# # Flowchart of for Loop
# # Flowchart of for Loop in Python programming
# # Flowchart of for Loop in Python
# # Example: Python for Loop
# # Program to find the sum of all numbers stored in a list

# # List of numbers
# # numbers = ["6", 5, 3, 8, 4, 2, 5, 4, 11]
# # # variable to store the sum
# # sum = 0
# # # iterate over the list
# # for val in numbers:
# #     sum = sum+int(val)

# # print("The sum is", sum)

# # Python Program to find the L.C.M. of two input number

# # def compute_lcm(x, y):

# #    # choose the greater number
# #    if x > y:
# #        greater = x
# #    else:
# #        greater = y

# #    while(True):
# #        if((greater % x == 0) and (greater % y == 0)):
# #            lcm = greater
# #            break
# #        greater += 1

# #    return lcm

# # num1 = 2
# # num2 = 3

# # print("The L.C.M. is", compute_lcm(num1, num2))


# # Program to add two matrices using nested loop

# # X = [[12,7,3],
# #     [4 ,5,6],
# #     [7 ,8,9]]

# # Y = [[5,8,1],
# #     [6,7,3],
# #     [4,5,9]]

# # result = [[0,0,0],
# #          [0,0,0],
# #          [0,0,0]]

# # # iterate through rows
# # for i in range(len(X)):
# #    # iterate through columns
# #    for j in range(len(X[0])):
# #        result[i][j] = X[i][j] + Y[i][j]

# # for r in result:
# #    print(r)


# # rows = int(input("Enter number of rows: "))

# # for i in range(rows, 1, -1):
# #     for space in range(0, rows-i):
# #         print("  ", end="")
# #     for j in range(i, 2*i-1):
# #         print("* ", end="")
# #     for j in range(1, i-1):
# #         print("* ", end="")
# #     print()

# # Python program to find the largest number among the three input numbers

# # change the values of num1, num2 and num3
# # for a different result
# # num1 = 106
# # num2 = 106
# # num3 = 106

# # # uncomment following lines to take three numbers from user
# # #num1 = float(input("Enter first number: "))
# # #num2 = float(input("Enter second number: "))
# # #num3 = float(input("Enter third number: "))

# # if (num1 >= num2) and (num1 >= num3):
# #    largest = num1
# # elif (num2 >= num1) and (num2 >= num3):
# #    largest = num2
# # else:
# #    largest = num3

# # print("The largest number is", largest)


# # class Parrot:

# #     # class attribute
# #     species = "bird"

# #     # instance attribute
# #     def __init__(self, name, age):
# #         self.name = name
# #         self.age = age

# # # instantiate the Parrot class
# # blu = Parrot("Blu", 10)
# # woo = Parrot("Woo", 15)

# # # access the class attributes
# # print("Blu is a {}".format(blu.__class__.species))
# # print("Woo is also a {}".format(woo.__class__.species))

# # # access the instance attributes
# # print("{} is {} years old".format( blu.name, blu.age))
# # print("{} is {} years old".format( woo.name, woo.age))


# # class Parrot:
    
# #     # instance attributes
# #     def __init__(self, name, age):
# #         self.name = name
# #         self.age = age
    
# #     # instance method
# #     def sing(self, song):
# #         return "{} sings {}".format(self.name, song)

# #     def dance(self):
# #         return "{} is now dancing".format(self.name)

# # # instantiate the object
# # blu = Parrot("Blu", 10)

# # # call our instance methods
# # print(blu.sing("'Happy'"))
# # print(blu.dance())


# # f = open("test.txt", encoding = 'utf-8')
# # # perform file operations
# # f.close()

# # try:
# # 	f = open("test.txt", encoding = 'utf-8')
# #    	# perform file operations
# # finally:
# # 	f.close()

# # with open("test.txt",'w',encoding = 'utf-8') as f:
# #    f.write("my first file\n")
# #    f.write("This file\n\n")
# #    f.write("contains three lines\n")


# # f = open("test.txt",'r',encoding = 'utf-8')


# # print(f.readline())
# # print(f.tell())


# # for x in f:
# # 	print(x, end = '')

# # print(f.readlines())
# # print(f.readlines())
# # print(f.readlines())
# # print(f.readlines())


# # import os

# # print(os.getcwdb())
# # print(os.listdir())

# # rows = int(input("Enter number of rows: "))

# # for i in range(rows+1):
# #     for j in range(rows,0):
# #         print("* ", end="")
# #     print("\n")

# # for x in range(5, -1, -1):
# # 	print(x)

# # import datetime

# # from datetime import date


# # from datetime import time

# # from datetime import datetime

# # from datetime import datetime, date

# # from datetime import timedelta
# # import pytz
# # import time


# # import threading 
  
# # def print_hello_three_times():
# #   for i in range(3):
# #     print("Hello")
  
# # def print_hi_three_times(): 
# #     for i in range(3): 
# #       print("Hi") 

# # t1 = threading.Thread(target=print_hello_three_times)  
# # t2 = threading.Thread(target=print_hi_three_times)  

# # t1.start()
# # t2.start()



# # print("Printed immediately.")
# # time.sleep(2.4)
# # print("Printed after 2.4 seconds.")


# # while True:
# #   localtime = time.localtime()
# #   result = time.strftime("%I:%M:%S %p", localtime)
# #   print(result)
# #   time.sleep(1)


# # while True:
# #   localtime = time.localtime()
# #   result = time.strftime("%I:%M:%S %p", localtime)
# #   print(result, end="", flush=True)
# #   print("\r", end="", flush=True)
# #   time.sleep(1)



# # #Get Current Date and Time
# # datetime_object = datetime.datetime.now()

# # print(datetime_object)

# # # Get Current Date

# # print(datetime.date.today())

# # print(dir(datetime))

# # print(datetime.date(2019, 4, 13))

# # a = date(2019, 4, 13)
# # print(a)

# # today = date.today()

# # print("Current date =", today)


# # timestamp = date.fromtimestamp(1326244364)
# # print("Date =", timestamp)


# # # date object of today's date
# # today = date.today() 
# # print(today)
# # print("Current year:", today.year)
# # print("Current month:", today.month)
# # print("Current day:", today.day)


# # # time(hour = 0, minute = 0, second = 0)
# # a = time()
# # print("a =", a)

# # # time(hour, minute and second)
# # b = time(11, 34, 56)
# # print("b =", b)


# # # time(hour, minute and second)
# # c = time(hour = 11, minute = 34, second = 56)
# # print("c =", c)

# # # time(hour, minute, second, microsecond)
# # d = time(11, 34, 56, 234566)
# # print("d =", d)


# # a = time(11, 34, 56)

# # print("hour =", a.hour)
# # print("minute =", a.minute)
# # print("second =", a.second)
# # print("microsecond =", a.microsecond)

# # #datetime(year, month, day)
# # a = datetime(2018, 11, 28)
# # print(a)

# # # datetime(year, month, day, hour, minute, second, microsecond)
# # b = datetime(2017, 11, 28, 23, 55, 59, 342380)
# # print(b)

# # a = datetime(2017, 11, 28, 23, 55, 59, 342380)
# # print("year =", a.year)
# # print("month =", a.month)
# # print("hour =", a.hour)
# # print("minute =", a.minute)
# # print("timestamp =", a.timestamp())

# # t1 = date(year = 2018, month = 7, day = 12)
# # t2 = date(year = 2017, month = 12, day = 23)
# # t3 = t1 - t2
# # print("t3 =", t3)

# # t4 = datetime(year = 2018, month = 7, day = 12, hour = 7, minute = 9, second = 33)
# # t5 = datetime(year = 2019, month = 6, day = 10, hour = 5, minute = 55, second = 13)
# # t6 = t4 - t5
# # print("t6 =", t6)


# # print("type of t3 =", type(t3)) 
# # print("type of t6 =", type(t6))


# # t1 = timedelta(weeks = 2, days = 5, hours = 1, seconds = 33)
# # t2 = timedelta(days = 4, hours = 11, minutes = 4, seconds = 54)
# # t3 = t1 - t2

# # print("t3 =", t3)


# # t1 = timedelta(seconds = 33)
# # t2 = timedelta(seconds = 54)
# # t3 = t1 - t2

# # print("t3 =", t3)
# # print("t3 =", abs(t3))

# # from datetime import timedelta

# # t = timedelta(days = 5, hours = 1, seconds = 33, microseconds = 233423)
# # print("total seconds =", t.total_seconds())


# # now = datetime.now()

# # t = now.strftime("%H:%M:%S")
# # print("time:", t)

# # s1 = now.strftime("%m/%d/%Y, %H:%M:%S")
# # # mm/dd/YY H:M:S format
# # print("s1:", s1)

# # s2 = now.strftime("%d/%m/%Y, %H:%M:%S")
# # # dd/mm/YY H:M:S format
# # print("s2:", s2)


# # date_string = "21 June, 2018"
# # print("date_string =", date_string)

# # date_object = datetime.strptime(date_string, "%d %B, %Y")
# # print("date_object =", date_object)

# # local = datetime.now()
# # print("Local:", local.strftime("%m/%d/%Y, %H:%M:%S"))


# # tz_NY = pytz.timezone('America/New_York') 
# # datetime_NY = datetime.now(tz_NY)
# # print("NY:", datetime_NY.strftime("%m/%d/%Y, %H:%M:%S"))

# # tz_London = pytz.timezone('Europe/London')
# # datetime_London = datetime.now(tz_London)
# # print("London:", datetime_London.strftime("%m/%d/%Y, %H:%M:%S"))

# # now = datetime.now() # current date and time

# # year = now.strftime("%Y")
# # print("year:", year)

# # month = now.strftime("%m")
# # print("month:", month)

# # day = now.strftime("%d")
# # print("day:", day)

# # time = now.strftime("%H:%M:%S")
# # print("time:", time)

# # date_time = now.strftime("%m/%d/%Y, %H:%M:%S")
# # print("date and time:",date_time)


# # timestamp = 1528797322
# # date_time = datetime.fromtimestamp(timestamp)

# # print("Date time object:", date_time)

# # d = date_time.strftime("%m/%d/%Y, %H:%M:%S")
# # print("Output 2:", d)	

# # d = date_time.strftime("%d %b, %Y")
# # print("Output 3:", d)

# # d = date_time.strftime("%d %B, %Y")
# # print("Output 4:", d)

# # d = date_time.strftime("%I%p")
# # print("Output 5:", d)


# # timestamp = 1528797322
# # date_time = datetime.fromtimestamp(timestamp)

# # d = date_time.strftime("%c")
# # print("Output 1:", d)	

# # d = date_time.strftime("%x")
# # print("Output 2:", d)

# # d = date_time.strftime("%X")
# # print("Output 3:", d)

# # def my_func():
# # 	x = 10
# # 	print("Value inside function:",x)

# # x = 20
# # my_func()
# # print("Value outside function:",x)

# # def greet(*names):
# #     """This function greets all
# #     the person in the names tuple."""

# #     # names is a tuple with arguments
# #     for name in names:
# #         print("Hello", name)


# # greet("Monica", "Luke", "Steve", "John")

# # def factorial(x):
# #     """This is a recursive function
# #     to find the factorial of an integer"""

# #     if x == 1:
# #         return 1
# #     else:
# #         return (x * factorial(x-1))


# # num = 3
# # print("The factorial of", num, "is", factorial(num))

# # Program to show the use of lambda functions
# # double = lambda x: x * 2

# # print(double(5))

# # # Program to filter out only the even items from a list
# # my_list = [1, 5, 4, 6, 8, 11, 3, 12]

# # new_list = list(filter(lambda x: (x%2 == 0) , my_list))

# # print(new_list)

# # x = "global"

# # def foo():
# #     print("x inside:", x)


# # foo()
# # print("x outside:", x)


# # x = "global"

# # def foo():
# #     x = x * 2
# #     print(x)

# # foo()

# # def foo():
# #     y = "local"


# # foo()
# # print(y)


# # def outer():
# #     x = "local"

# #     def inner():
# #         nonlocal x
# #         x = "nonlocal"
# #         print("inner:", x)

# #     inner()
# #     print("outer:", x)


# # outer()

# # c = 1 # global variable

# # def add():
# #     print(c)

# # add()

# # import config
# # import update

# # print(config.a)
# # print(config.b)

# # import all names from the standard module math

# # from math import *
# # print("The value of pi is", pi)

# # import Game.Level.start
# # Game.Level.start.select_difficulty(2)


# # class SalaryNotInRangeError(Exception):

# #     def __init__(self, salary, message="Salary is not in (5000, 15000) range"):
# #         self.salary = salary
# #         self.message = message
# #         super().__init__(self.message)

# #     def __str__(self):
# #         return f'{self.salary} -> {self.message}'


# # salary = int(input("Enter salary amount: "))
# # if not 5000 < salary < 15000:
# #     raise SalaryNotInRangeError(salary)

# # str1 = 'cAr'
# # str2 = 'Rac'

# # str1 = str1.lower()
# # str2 = str2.lower()

# # if len(str1) == len(str2):
	
# # 	str1_sorted = sorted(str1)
# # 	str2_sorted = sorted(str2)

# # 	if str1_sorted==str2_sorted:
# # 		print("Yes")
# # 	else:
# # 		print("No")

# # else:
# # 	print("Not")
# # str1 = 'cAr'
# # str2 = 'aDc'

# # str1_lower = str1.lower()
# # str2_lower = str2.lower()

# # print(sorted(str1_lower))
# # print(sorted(str2_lower))

# # print(str1_lower)
# # print(str2_lower)

# # print(len(str1_lower))

# # if len(str1_lower) == len(str2_lower):
# # 	if sorted(str1_lower)==sorted(str2_lower):
# # 		print("Yes")
# # 	else:
# # 		print("No")
# # else:
# # 	print("No")


# # for x in range(1,11):
# # 	if x%2!=0:
# # 		print(x)

# import smtplib, ssl

# # port = 465  # For SSL
# # password = input("Type your password and press enter: ")

# # # Create a secure SSL context
# # context = ssl.create_default_context()

# # with smtplib.SMTP_SSL("smtp.gmail.com", port, context=context) as server:
# #     server.login("dream.sharvan@gmail.com", password)
# #     # TODO: Send email here

# # sender_email = "my@gmail.com"
# # receiver_email = "your@gmail.com"
# # message = """\
# # Subject: Hi there

# # This message is sent from Python."""

# # Send email here

# port = 465  # For SSL
# password = input("Type your password and press enter: ")

# # Create a secure SSL context
# context = ssl.create_default_context()



# sender_email = "dream.sharvan@gmail.com"
# receiver_email = "dream.sharvan@gmail.com"
# message = """\
# Subject: Hi there

# This message is sent from Python."""

# # Send email here

# if True:
#    smtpObj = with smtplib.SMTP_SSL("smtp.gmail.com", port, context=context) as server:
#     server.login("dream.sharvan@gmail.com", password)
#    smtpObj.sendmail(sender, receivers, message)         
#    print ("Successfully sent email")
# else:
#    print ("Error: unable to send email")

#Three lines to make our compiler able to draw:
# import sys
# import matplotlib
# matplotlib.use('Agg')

# import matplotlib.pyplot as plt
# import numpy as np

# ypoints = np.array([3, 8, 1, 10])

# plt.plot(ypoints, marker = 'o')
# plt.show()

# #Two  lines to make our compiler able to draw:
# plt.savefig(sys.stdout.buffer)
# sys.stdout.flush()


# import matplotlib.pyplot as plt
# import numpy as np

# xpoints = np.array([100, 60])
# ypoints = np.array([0, 250])

# plt.plot(xpoints, ypoints)
# plt.show()

#Three lines to make our compiler able to draw:
# import sys
# import matplotlib
# matplotlib.use('Agg')

# import matplotlib.pyplot as plt
# import numpy as np

# ypoints = np.array([3, 8, 1, 10])

# plt.plot(ypoints, marker = '*')
# plt.show()

# #Two  lines to make our compiler able to draw:
# plt.savefig(sys.stdout.buffer)
# sys.stdout.flush()


# import numpy
# import matplotlib.pyplot as plt

# x = numpy.random.normal(5.0, 1.0, 100000)

# plt.hist(x, 100)
# plt.show()


# num1 = input("Enter The First Number: ")
# num2 = input("Enter The Second Number: ")
# e1 = e2 = ''
# if num1=='':
# 	print("Num1 Can Not be blank")
# else:
# 	e1 = True

# if num1=='':
# 	print("Num2 Can Not be blank")
# else:
# 	e2= True

# if e1 and e2 == True:
# 	print("Valid")

# a = 10
# b = 20

# print("The value of a is {} and b is {}".format(a,b))


myorder = "I have a {}, it is a {}."

print(myorder.format( "Ford", "Mustang"))
